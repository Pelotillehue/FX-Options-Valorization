"""
Pipeline diario de valorizacion de Opciones FX (USD/CLP).

Entradas (los 2 adjuntos del correo diario):
  1. Grid Vol Bloomberg Creasys.xlsx           -> datos de mercado
  2. Operaciones Diarias ... al YYYY_MM_DD.xlsx -> cartera / altas / vencimientos

Reconstruye la cartera vigente del dia como {inicio <= fecha_proceso < vencimiento}
(auto-consistente, sin estado acumulado), valoriza cada operacion, liquida las
que vencen en la fecha de proceso usando el fixing USDOBS, y genera:
  - MtM_YYYYMMDD.xlsx                (cartera vigente valorizada)
  - MtM_YYYYMMDD_vencimientos.xlsx   (operaciones que vencen + flujo de pago)
  - resumen_YYYYMMDD.txt             (conteos tipo correo)

Uso:
  python daily_process.py <grid.xlsx> <operaciones.xlsx> [YYYY-MM-DD] [dir_salida]
"""
import sys, os, re
from datetime import date, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

from dateparse import parse_fecha
from load_market import load_market
from fx_options_valuation import Operacion, liquidar_vencimiento


# --------------------------------------------------------------------------
def leer_operaciones(path):
    """Lee la hoja 'Operaciones Diarias' -> lista de dicts con todos los campos."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Operaciones Diarias"]
    ops, avisos = [], []
    for r in range(2, ws.max_row + 1):
        folio = ws.cell(r, 1).value
        if not folio:
            continue
        inicio = parse_fecha(ws.cell(r, 2).value)
        venc = parse_fecha(ws.cell(r, 3).value)
        if inicio is None or venc is None:
            avisos.append(f"fila {r} ({folio}): fecha no reconocida "
                          f"inicio={ws.cell(r,2).value!r} venc={ws.cell(r,3).value!r}")
            continue
        ops.append(dict(
            folio=folio, inicio=inicio, venc=venc,
            contraparte=ws.cell(r, 6).value, lado=ws.cell(r, 7).value,
            tipo=ws.cell(r, 8).value, modalidad=ws.cell(r, 9).value,
            entrega=ws.cell(r, 10).value, moneda=ws.cell(r, 11).value,
            nominal=ws.cell(r, 12).value, strike=ws.cell(r, 13).value,
            par=ws.cell(r, 5).value,
            prima_clp_usd=ws.cell(r, 15).value, prima_pct=ws.cell(r, 17).value,
            prima_clp=ws.cell(r, 16).value, prima_usd=ws.cell(r, 18).value,
            margen_clp=ws.cell(r, 22).value, margen_usd=ws.cell(r, 24).value,
            usdobs=ws.cell(r, 25).value, itm_src=ws.cell(r, 26).value,
        ))
    wb.close()
    return ops, avisos


def clasificar(ops, fecha_proceso):
    """vigentes (a valorizar) y vencimientos (a liquidar en la fecha de proceso)."""
    vigentes = [o for o in ops if o["inicio"] <= fecha_proceso < o["venc"]]
    vencen = [o for o in ops if o["venc"] == fecha_proceso]
    altas = [o for o in ops if o["inicio"] == fecha_proceso]
    return vigentes, vencen, altas


def _fx_prima(op, universo):
    """Tipo de cambio pagado en la prima = Margen Comercial CLP / Margen Comercial USD.
    Si esta pata no tiene margen, hereda el del par (misma fecha/nominal/strike)."""
    mc, mu = op.get("margen_clp"), op.get("margen_usd")
    if mu not in (None, 0):
        return mc / mu
    for o in universo:
        if (o is not op and o["inicio"] == op["inicio"] and o["venc"] == op["venc"]
                and o["nominal"] == op["nominal"] and o["strike"] == op["strike"]
                and o.get("margen_usd") not in (None, 0)):
            return o["margen_clp"] / o["margen_usd"]
    return None


def _clave_calce(op):
    """Identifica el grupo de operaciones que se calzan entre si."""
    return (op["inicio"], op["venc"], op["strike"], (op["tipo"] or "").strip().upper())


def _fx_por_calce(ops):
    """Tipo de cambio implicito de las primas, calculado POR CALCE: para cada grupo
    de operaciones calzadas, suma del margen comercial en CLP dividida por la suma
    neta de primas en USD del grupo. Sirve para calces 1 a 1, N a 1 y N a M (agrega
    las primas de todas las patas del grupo)."""
    acc = {}
    for o in ops:
        if not _es_prima_pct(o):
            continue
        k = _clave_calce(o)
        num, den = acc.get(k, (0.0, 0.0))
        acc[k] = (num + (o.get("margen_clp") or 0), den + (o.get("prima_usd") or 0))
    return {k: n / d for k, (n, d) in acc.items() if d}


def _es_prima_pct(op):
    """True si la prima se pacto como % del nominal USD (no como CLP por USD)."""
    return (op.get("prima_clp") in (None, 0)) and (
        op.get("prima_pct") not in (None, 0) or op.get("prima_usd") not in (None, 0))


def _enriquecer_pl(vigentes, resultados):
    """Completa P&L No Realizado en CLP y USD segun el tipo de prima de cada operacion."""
    fx_calce = _fx_por_calce(vigentes)  
    for op, res in zip(vigentes, resultados):
        mtm, spot = res["mtm"], res["spot"]
        pclp, pusd = op.get("prima_clp"), op.get("prima_usd")
        if _es_prima_pct(op):                         # prima como % del nominal USD
            fx = fx_calce.get(_clave_calce(op)) or _fx_prima(op, vigentes) or spot
            pl_usd = (pusd + mtm / spot) if pusd not in (None, 0) else 0.0
            res["prima_clp_disp"] = None
            res["prima_usd_disp"] = pusd
            res["pl_usd"] = pl_usd
            res["pl_clp"] = pl_usd * fx
            res["fx_prima"] = fx
        else:                                         # prima en CLP por USD (o sin prima)
            res["prima_clp_disp"] = pclp
            res["prima_usd_disp"] = pusd
            res["pl_usd"] = None
            res["pl_clp"] = (pclp + mtm) if pclp not in (None, 0) else 0.0
    return resultados


def _op_obj(o):
    return Operacion(folio=o["folio"], fecha_inicio=o["inicio"],
                     fecha_vencimiento=o["venc"], tipo=o["tipo"], lado=o["lado"],
                     nominal=o["nominal"] or 0, strike=o["strike"] or 0,
                     prima_clp=o["prima_clp"] or 0)


# --------------------------------------------------------------------------
# Estilos Excel
# --------------------------------------------------------------------------
_HDR_FILL = PatternFill("solid", fgColor="366092")
_HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
_TIT_FONT = Font(bold=True, size=13, color="366092")
_THIN = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_MONEY = '#,##0;[Red]-#,##0'
_NUM2 = '#,##0.00'
_NUM4 = '#,##0.0000'

# Ruta del logo (opcional). Deja tu imagen en assets/logo.png (o .jpg).
LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "logo.png")


# Tablas y titulos comienzan en esta columna (1=A, 2=B, ...); la(s) columna(s)
# anteriores quedan vacias, a modo de margen.
TABLE_COL0 = 2

# Ancho de la columna A: margen vacio a la izquierda de la tabla
COL_A_WIDTH = 3

# Ancho del logo en pixeles (independiente del ancho de la columna A)
LOGO_WIDTH_PX = 131


def _ancho_col_px(width):
    """Convierte ancho de columna Excel (unidades) a pixeles (fuente Calibri 11)."""
    return int(round(width * 7 + 5))


def _insertar_logo(ws, anchor="A1", ancho_px=None, alto_px=40):
    """Inserta el logo escalado para caber en ancho_px x alto_px (mantiene proporcion).
    Si el archivo no existe (o falta Pillow), no hace nada."""
    if not os.path.exists(LOGO_PATH):
        return
    if ancho_px is None:
        ancho_px = LOGO_WIDTH_PX
    try:
        from openpyxl.drawing.image import Image as XLImage
        img = XLImage(LOGO_PATH)
        w0 = img.width or ancho_px
        h0 = img.height or alto_px
        ratio = min(ancho_px / w0, alto_px / h0)              # el limite mas restrictivo manda
        img.width = int(w0 * ratio)
        img.height = int(h0 * ratio)
        ws.add_image(img, anchor)
        alto_pts = img.height * 0.75                          # px -> puntos (aprox)
        actual = ws.row_dimensions[1].height or 0
        ws.row_dimensions[1].height = max(actual, alto_pts)
    except Exception:
        pass                                                  # si falta Pillow u otro, no romper


def _estilo_encabezado(ws, fila, ncols, col0=TABLE_COL0):
    for c in range(col0, col0 + ncols):
        cell = ws.cell(fila, c)
        cell.fill = _HDR_FILL; cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER


_MESES_ABR = ["ene","feb","mar","abr","may","jun","jul","ago","sep","oct","nov","dic"]

# Formato moneda: "$1.234" y negativos "-$1.234" (sin color rojo)
_MON  = r'_ "$"* #,##0_ ;_ "$"* \-#,##0_ ;_ "$"* "-"_ ;_ @_ '   # contabilidad: $ fijo a la izquierda
_MON2 = r'_ "$"* #,##0.00_ ;_ "$"* \-#,##0.00_ ;_ "$"* "-"_ ;_ @_ '
_PCT7 = '0.0000000%'          # porcentaje con 7 decimales

# Columnas que van centradas (el resto va a la derecha en formato moneda)
_CENTRADAS = {
    "Folio", "Fecha Inicio", "Fecha Vencimiento", "Plazo (días)",
    "Fecha Proceso", "Par de Monedas", "Contraparte", "Lado Nevasa", "Tipo",
    "Tipo  Modalidad", "Tipo Entrega", "Moneda Principal", "Spot",
    "Volatilidad", "Forward", "Strike", "Spot (Fixing)", "Prima (% Monto USD)",
}
_BAND_CLARA = PatternFill("solid", fgColor="DCE6F1")   # filas impares (1a, 3a, ...)
_BAND_OSCURA = PatternFill("solid", fgColor="B8CCE4")  # filas pares (2a, 4a, ...)
_TOTAL_FILL = PatternFill("solid", fgColor="FFFFFF")


def _fecha_corta(d):
    return f"{d.day:02d}-{_MESES_ABR[d.month - 1]}-{d:%y}"


def _volcar_tabla(ws, colspecs, filas, hr, tot_idx=None, col0=TABLE_COL0):
    """Escribe encabezado (fila hr) y filas de datos segun colspecs=[(titulo, fmt)].
    col0 es la columna (1=A, 2=B, ...) donde comienza la tabla."""
    for i, (h, _fmt) in enumerate(colspecs, col0):
        ws.cell(hr, i, h)
    _estilo_encabezado(ws, hr, len(colspecs), col0=col0)
    for k, vals in enumerate(filas):
        row = hr + 1 + k
        for i, (h, fmt) in enumerate(colspecs, col0):
            v = vals[i - col0]
            c = ws.cell(row, i, v)
            c.border = _BORDER
            c.alignment = Alignment(horizontal="center" if h in _CENTRADAS else "right")
            if fmt is not None and v is not None:
                c.number_format = fmt
            c.fill = _BAND_OSCURA if k % 2 == 1 else _BAND_CLARA
            if tot_idx is not None and (i - col0) in tot_idx and isinstance(v, (int, float)):
                tot_idx[i - col0] += v
    if tot_idx is not None:
        tr = hr + 1 + len(filas)
        #tc = ws.cell(tr, col0, "TOTAL"); tc.font = Font(bold=True); tc.fill = _TOTAL_FILL
        for idx, total in tot_idx.items():
            c = ws.cell(tr, idx + col0, total)
            c.font = Font(bold=True); c.number_format = _MON; c.fill = _TOTAL_FILL
            c.alignment = Alignment(horizontal="right")
        return tr
    return hr + len(filas)


def escribir_mtm(vigentes, resultados, md, path, fecha_proceso):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "MtM"
    ws.column_dimensions["A"].width = COL_A_WIDTH
    _insertar_logo(ws, "B1")
    ws["D1"] = f"Mark-to-Market Opciones FX al {fecha_proceso:%d-%b-%Y}"
    ws["D1"].font = _TIT_FONT
    ws["D1"].alignment = Alignment(vertical="center")
    ws["D2"] = f"Spot USDCLP: {md.spot:,.2f} Operaciones vigentes: {len(vigentes)}"
    ws["D2"].font = Font(italic=True, size=9, color="555555")

    F2, MONEY, NUM2, PCT = "dd-mmm-yy", _MONEY, _NUM2, "0.00%"
    mtm_hdr = f"MtM al {_fecha_corta(fecha_proceso)}"
    colspecs = [
        ("Folio", None), ("Fecha Inicio", F2), ("Fecha Vencimiento", F2),
        ("Plazo (días)", "0"), ("Fecha Proceso", F2), ("Par de Monedas", None),
        ("Contraparte", None), ("Lado Nevasa", None), ("Tipo", None),
        ("Tipo  Modalidad", None), ("Tipo Entrega", None), ("Moneda Principal", None),
        ("Spot", NUM2), ("Volatilidad", _NUM4), ("Forward", _NUM4),
        ("Monto Principal", _MON), ("Strike", NUM2), ("Prima (CLP por USD)", _MON2),
        ("Prima (% Monto USD)", _PCT7), ("Prima (CLP)", _MON), ("Prima (USD)", _MON2),
        (mtm_hdr, _MON), ("P&L No Realizado Nevasa (CLP)", _MON),
        ("P&L No Realizado Nevasa (USD)", _MON2),
    ]
    filas = []
    for o, res in zip(vigentes, resultados):
        plazo = (o["venc"] - o["inicio"]).days
        filas.append([
            o["folio"], o["inicio"], o["venc"], plazo, fecha_proceso, o["par"],
            o["contraparte"], o["lado"], o["tipo"], o["modalidad"], o["entrega"],
            o["moneda"], md.spot, res["vol"], res["forward"], o["nominal"], o["strike"],
            o["prima_clp_usd"], o["prima_pct"], res["prima_clp_disp"], res["prima_usd_disp"],
            res["mtm"], res["pl_clp"], res["pl_usd"],
        ])
    tot_idx = {19: 0.0, 21: 0.0, 22: 0.0}   # Prima (CLP), MtM, P&L CLP
    hr = 4                                                              # ← NUEVO
    _volcar_tabla(ws, colspecs, filas, hr=hr, tot_idx=tot_idx)          # ← usa hr en vez de 4
    # nota con el tipo de cambio usado, en el P&L CLP de las operaciones con prima %
    col_pl = next(i for i, (h, _f) in enumerate(colspecs, TABLE_COL0)    # ← NUEVO (bloque completo)
                  if h == "P&L No Realizado Nevasa (CLP)")
    for k, res in enumerate(resultados):
        fx = res.get("fx_prima")
        if fx:
            ws.cell(hr + 1 + k, col_pl).comment = Comment(
                f"P&L en CLP con tipo de cambio {fx:,.2f}", "Fabián Mancilla")
    _ajustar_anchos(ws, colspecs, filas)
    ws.column_dimensions["A"].width = COL_A_WIDTH
    ws.freeze_panes = "A5"
    wb.save(path)


def escribir_vencimientos(vencen, liqs, path, fecha_proceso):
    """Una tabla si todos los vencimientos son del mismo tipo de prima; dos tablas
    (primero CLP por USD, luego % del nominal USD) si hay de ambos tipos."""
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Vencimientos"
    ws.column_dimensions["A"].width = COL_A_WIDTH
    _insertar_logo(ws, "B1")
    ws["D1"] = f"Vencimientos al {fecha_proceso:%d-%b-%Y}"
    ws["D1"].font = _TIT_FONT
    ws["D1"].alignment = Alignment(vertical="center")

    F2, NUM2 = "dd-mmm-yy", _NUM2
    payoff_hdr = f"Payoff {_fecha_corta(fecha_proceso)} (CLP)"
    anchos = {}
    fx_calce = _fx_por_calce(vencen)
    pares = list(zip(vencen, liqs))
    normales = [(o, l) for o, l in pares if not _es_prima_pct(o)]
    pct = [(o, l) for o, l in pares if _es_prima_pct(o)]

    def _tabla(hr, grupo, tipo):
        if tipo == "clp":
            colQ, colR = ("Prima (CLP por USD)", _MON2), ("Prima (CLP)", _MON)
        else:
            colQ, colR = ("Prima (% Monto USD)", _PCT7), ("Prima (USD)", _MON2)
        colspecs = [
            ("Folio", None), ("Fecha Inicio", F2), ("Fecha Vencimiento", F2),
            ("Plazo (días)", "0"), ("Fecha Proceso", F2), ("Par de Monedas", None),
            ("Contraparte", None), ("Lado Nevasa", None), ("Tipo", None),
            ("Tipo  Modalidad", None), ("Tipo Entrega", None), ("Moneda Principal", None),
            ("Spot (Fixing)", NUM2), ("Monto Principal", _MON), ("Strike", NUM2),
            colQ, colR, (payoff_hdr, _MON), ("P&L Nevasa (CLP)", _MON),
        ]
        filas = []
        fxs = []    # fx usado por fila (solo tabla prima pagan % de nominal en USD)
        for o, liq in grupo:
            plazo = (o["venc"] - o["inicio"]).days
            payoff = liq["flujo_pago_clp"]
            if tipo == "clp":
                pq, pr = o["prima_clp_usd"], (o["prima_clp"] or 0)
                pl = (o["prima_clp"] or 0) + payoff
                fxs.append(None)
            else:
                #fx = _fx_prima(o, vencen) or liq["fixing"]
                fx = fx_calce.get(_clave_calce(o)) or _fx_prima(o, vencen) or liq["fixing"]
                pq, pr = o["prima_pct"], o["prima_usd"]
                pl = (o["prima_usd"] or 0) * fx + payoff
                fxs.append(fx)
            filas.append([
                o["folio"], o["inicio"], o["venc"], plazo, fecha_proceso, o["par"],
                o["contraparte"], o["lado"], o["tipo"], o["modalidad"], o["entrega"],
                o["moneda"], liq["fixing"], o["nominal"], o["strike"], pq, pr, payoff, pl,
            ])
        tot_idx = {16: 0.0, 17: 0.0, 18: 0.0}   # Prima (R), Payoff, P&L Nevasa
        last = _volcar_tabla(ws, colspecs, filas, hr, tot_idx)
        # nota con el tipo de cambio usado, en el P&L Nevasa de las operaciones con prima como % del nominal USD
        col_pl = TABLE_COL0 + len(colspecs) - 1    # ultima columna = P&L Nevasa
        for k, fx in enumerate(fxs):
            if fx:
                ws.cell(hr + 1 + k, col_pl).comment = Comment(
                    f"P&L en CLP con tipo de cambio {fx:,.2f}", "Fabián Mancilla")
        _ajustar_anchos(ws, colspecs, filas, anchos)
        return last, len(colspecs)

    hr = 4
    ncols = 19
    if normales:
        last, ncols = _tabla(hr, normales, "clp")
        hr = last + 3                       # 2 filas en blanco entre tablas
    if pct:
        last, ncols = _tabla(hr, pct, "pct")
    
    # ensanchar 50% las ultimas 3 columnas: Prima, Payoff y P&L Nevasa
    ultima = TABLE_COL0 + ncols - 1
    for c in (ultima - 2, ultima - 1, ultima):
        letra = get_column_letter(c)
        w = ws.column_dimensions[letra].width or 10
        ws.column_dimensions[letra].width = round(w * 1.5, 1)

    ws.column_dimensions["A"].width = COL_A_WIDTH
    ws.freeze_panes = "A5"
    wb.save(path)


def _decimales(fmt):
    """Decimales que muestra un formato de numero de Excel."""
    if not fmt or "." not in fmt:
        return 0
    seccion = fmt.split(";")[0]
    if "." not in seccion:
        return 0
    return sum(1 for ch in seccion.split(".", 1)[1] if ch in "0#")


def _largo_visible(v, fmt):
    """Largo del texto tal como se vera en pantalla (no del valor crudo)."""
    if v is None:
        return 0
    if isinstance(v, (datetime, date)):
        return 9                                   # dd-mmm-yy
    if isinstance(v, (int, float)):
        dec = _decimales(fmt)
        if fmt and "%" in fmt:
            return len(f"{v * 100:,.{dec}f}") + 1
        s = f"{v:,.{dec}f}"
        if fmt and "$" in fmt:
            s = "$" + s
        return len(s)
    return len(str(v))


def _ajustar_anchos(ws, colspecs, filas, acc=None, pad=2, max_w=45, col0=TABLE_COL0):
    """Ancho por columna: el mayor entre la palabra mas larga del titulo (los titulos
    se ajustan en varias lineas) y el dato mas largo ya formateado. 'acc' acumula
    anchos entre tablas de una misma hoja para no angostar la anterior."""
    if acc is None:
        acc = {}
    for i, (h, fmt) in enumerate(colspecs, col0):
        palabras = str(h).split()
        largo_hdr = max((len(p) for p in palabras), default=0)
        largo_val = max((_largo_visible(f[i - col0], fmt) for f in filas), default=0)
        w = min(max(largo_hdr, largo_val) + pad, max_w)
        acc[i] = max(acc.get(i, 0), w)
        ws.column_dimensions[get_column_letter(i)].width = acc[i]
    return acc


def _ancho_por_titulo(ws, colspecs, titulo, filas, max_w=45):
    """Ajusta el ancho de una columna (por su titulo) al contenido mas largo."""
    for i, (h, _f) in enumerate(colspecs, 1):
        if h == titulo:
            largo = max([len(str(titulo))] +
                        [len(str(f[i - 1])) for f in filas if f[i - 1] is not None])
            ws.column_dimensions[get_column_letter(i)].width = min(largo + 2, max_w)
            return


def _auto_width(ws, ncols):
    for c in range(1, ncols + 1):
        letra = get_column_letter(c)
        largo = 10
        for cell in ws[letra]:
            if cell.value is not None:
                largo = max(largo, len(str(cell.value)) + 2)
        ws.column_dimensions[letra].width = min(largo, 22)


# --------------------------------------------------------------------------
def procesar(grid_path, ops_path, fecha_proceso=None, out_dir="salidas"):
    os.makedirs(out_dir, exist_ok=True)
    if fecha_proceso is None:
        # fecha de proceso desde el nombre del archivo de cartera
        # ("... al YYYY_MM_DD.xlsx"); el grid queda como ultimo recurso.
        m = re.search(r"al (\d{4})_(\d{2})_(\d{2})", os.path.basename(ops_path))
        if m:
            fecha_proceso = date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    ops, avisos = leer_operaciones(ops_path)
    md = load_market(grid_path, fecha_proceso=fecha_proceso)
    avisos = list(avisos) + [f"datos de mercado: {a}" for a in (md.avisos or [])]
    fecha_proceso = md.fecha_proceso
    vigentes, vencen, altas = clasificar(ops, fecha_proceso)

    resultados = [_op_obj(o).valorizar(md) for o in vigentes]
    _enriquecer_pl(vigentes, resultados)
    sin_fixing = [o["folio"] for o in vencen
                  if not isinstance(o.get("usdobs"), (int, float))]
    if sin_fixing:
        raise ValueError("vencimientos sin fixing USDOBS en el archivo de "
                         "cartera: " + ", ".join(sin_fixing))
    liqs = [liquidar_vencimiento(_op_obj(o), o["usdobs"]) for o in vencen]

    tag = fecha_proceso.strftime("%Y%m%d")
    p_mtm = os.path.join(out_dir, f"MtM_{tag}.xlsx")
    p_ven = os.path.join(out_dir, f"MtM_{tag}_vencimientos.xlsx")
    p_res = os.path.join(out_dir, f"resumen_{tag}.txt")
    escribir_mtm(vigentes, resultados, md, p_mtm, fecha_proceso)
    escribir_vencimientos(vencen, liqs, p_ven, fecha_proceso)

    n_itm = sum(1 for l in liqs if l["itm"])
    tot_mtm = sum(r["mtm"] for r in resultados)
    tot_pl = sum(r["pl_clp"] for r in resultados)
    tot_flujo = sum(l["flujo_pago_clp"] for l in liqs)
    resumen = (
        f"Proceso de Valorización Opciones FX para el {fecha_proceso:%d-%m-%Y}\n"
        f"{'='*54}\n"
        f"Spot USDCLP: {md.spot:,.2f}\n\n"
        f"Hubo {len(altas)} nuevas operaciones.\n"
        + (f"Vencieron 0 operaciones.\n" if not vencen else
           f"Vencieron {len(vencen)} operaciones, ninguna ITM por lo que no hay "
           f"flujos de pago.\n" if n_itm == 0 else
           f"Vencieron {len(vencen)} operaciones, {n_itm} ITM por lo que hay "
           f"flujos de pago.\n")
        + f"Hay {len(vigentes)} operaciones vigentes.\n\n"
        f"MtM total cartera         : {tot_mtm:,.0f} CLP\n"
        f"P&L no realizado total    : {tot_pl:,.0f} CLP\n"
        f"Flujo de pago vencimientos: {tot_flujo:,.0f} CLP\n"
    )
    if altas:
        resumen += f"\nAltas del dia: {', '.join(o['folio'] for o in altas)}\n"
    if avisos:
        resumen += "\nAvisos:\n" + "\n".join("  - " + a for a in avisos) + "\n"
    with open(p_res, "w", encoding="utf-8") as f:
        f.write(resumen)
    return dict(fecha=fecha_proceso, vigentes=len(vigentes), vencen=len(vencen),
                itm=n_itm, altas=len(altas), tot_mtm=tot_mtm, tot_pl=tot_pl,
                tot_flujo=tot_flujo, avisos=avisos,
                archivos=[p_mtm, p_ven, p_res], resumen=resumen)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__); sys.exit(1)
    grid, opsf = sys.argv[1], sys.argv[2]
    # Los argumentos extra pueden venir en cualquier orden: si uno parsea como
    # fecha, es la fecha de proceso; si no, se toma como carpeta de salida.
    fp, outd = None, "salidas"
    for extra in sys.argv[3:]:
        d = parse_fecha(extra)
        if d is not None:
            fp = d
        else:
            outd = extra
    res = procesar(grid, opsf, fp, outd)
    print(res["resumen"])
    print("Archivos generados:")
    for a in res["archivos"]:
        print("  ", a)