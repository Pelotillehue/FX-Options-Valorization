"""Carga MarketData desde 'Grid Vol Bloomberg Creasys.xlsx' (reemplaza el notebook)."""
import openpyxl
from datetime import datetime
from fx_options_valuation import MarketData, STD_TENORS_DAYS

# etiqueta de tenor -> dias (para mapear filas de Mids BGN a dias estandar)
TENOR_DIAS = dict(zip(
    ["1D","1W","2W","3W","1M","2M","3M","4M","6M","9M","1Y","18M","2Y","3Y","4Y","5Y"],
    STD_TENORS_DAYS))

def _num(x):
    return x if isinstance(x, (int, float)) else None


def completar_fila_curva(spot, dias, bid, ask, mid_clp, fwd, mid_usd):
    """Completa datos faltantes de una fila de la curva:
    - Mid CLP vacio -> punto medio entre Bid y Ask.
    - Mid USD vacio -> tasa implicita desde FWD, Spot y la tasa CLP, usando
      FWD = Spot * DF_usd / DF_clp  (DF = 1/(1+r*d/360)):
          r_usd = ((Spot/FWD) * (1 + r_clp*d/360) - 1) * 360/d
    Devuelve (mid_clp, mid_usd); cualquiera puede quedar None si no es derivable."""
    mid_clp, mid_usd = _num(mid_clp), _num(mid_usd)
    bid, ask, fwd = _num(bid), _num(ask), _num(fwd)
    if mid_clp is None and bid is not None and ask is not None:
        mid_clp = (bid + ask) / 2
    if (mid_usd is None and fwd not in (None, 0) and mid_clp is not None
            and spot is not None and dias):
        mid_usd = ((spot / fwd) * (1 + mid_clp / 100 * dias / 360) - 1) * 360 / dias * 100
    return mid_clp, mid_usd


# Rango plausible (en %) para tasas de la curva; fuera de esto solo se ALERTA
# en el log (el dato se usa igual, sin filtrar).
TASA_MIN, TASA_MAX = -2.0, 25.0


def tasa_plausible(r):
    return r is not None and TASA_MIN <= r <= TASA_MAX


def load_market(grid_path, fecha_proceso=None):
    wb = openpyxl.load_workbook(grid_path, data_only=True)
    # ---- Curva cruda + spot desde 'FWD pts DEPOs' ----
    ws = wb["FWD pts DEPOs"]
    spot = ws["H28"].value            # 926.78
    fecha_hdr = ws["B28"].value       # fecha de proceso
    # Curvas asimetricas: la CLP usa todos sus nodos (hasta 20Y); la USD solo
    # los nodos con tasa implicita (hasta ~5Y), como en la hoja 'Interpolacion'.
    raw_dias, raw_clp, raw_usd = [], [], []
    clp_dias, clp_rates, usd_dias, usd_rates = [], [], [], []
    avisos = []
    r = 30
    while True:
        c = ws.cell(r, 3).value       # C = Dias
        if not isinstance(c, (int, float)):
            break
        g, i = completar_fila_curva(
            spot, float(c),
            ws.cell(r, 5).value,      # E = Bid
            ws.cell(r, 6).value,      # F = Ask
            ws.cell(r, 7).value,      # G = Mid CLP
            ws.cell(r, 8).value,      # H = FWD
            ws.cell(r, 9).value)      # I = Mid USD implicito
        if g is not None and not tasa_plausible(g):
            avisos.append(f"ALERTA curva CLP: nodo {int(c)}d con tasa implausible "
                          f"({g:.4g}%); revisar datos de mercado")
        if i is not None and not tasa_plausible(i):
            avisos.append(f"ALERTA curva USD: nodo {int(c)}d con tasa implausible "
                          f"({i:.4g}%); revisar datos de mercado")
        if g is not None:
            clp_dias.append(float(c)); clp_rates.append(float(g))
        if i is not None:
            usd_dias.append(float(c)); usd_rates.append(float(i))
        if g is not None and i is not None:
            raw_dias.append(float(c)); raw_clp.append(float(g)); raw_usd.append(float(i))
        r += 1
    # ---- Superficie de vol desde 'Mids BGN' (bloque 2, cols M:R) ----
    ws = wb["Mids BGN"]
    vol_tenor_dias, vol_smile = [], []
    for rr in range(3, 3 + 16):       # 1D .. 5Y
        term = ws.cell(rr, 13).value  # M = Term
        if term is None:
            continue
        term = str(term).strip()
        if term not in TENOR_DIAS:
            continue
        vatm = ws.cell(rr, 14).value  # N
        v25c = ws.cell(rr, 15).value  # O 25 Call
        v25p = ws.cell(rr, 16).value  # P 25 Put
        v10c = ws.cell(rr, 17).value  # Q 10 Call
        v10p = ws.cell(rr, 18).value  # R 10 Put
        vol_tenor_dias.append(TENOR_DIAS[term])
        vol_smile.append((float(v10p), float(v25p), float(vatm), float(v25c), float(v10c)))
    wb.close()
    fp = fecha_proceso or (fecha_hdr.date() if isinstance(fecha_hdr, datetime) else fecha_hdr)
    md = MarketData(fecha_proceso=fp, spot=float(spot),
                    raw_dias=raw_dias, raw_clp=raw_clp, raw_usd=raw_usd,
                    vol_tenor_dias=vol_tenor_dias, vol_smile=vol_smile,
                    clp_dias=clp_dias, clp_rates=clp_rates,
                    usd_dias=usd_dias, usd_rates=usd_rates,
                    avisos=avisos)
    return md.build()