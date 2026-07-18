from datetime import date
import openpyxl
import pytest
import daily_process as dp
from fx_options_valuation import liquidar_vencimiento


def op(folio, strike, lado, prima_usd=None, margen_clp=None, margen_usd=None,
       prima_clp=None, prima_clp_usd=None, prima_pct=None,
       inicio=date(2026, 4, 29), venc=date(2026, 5, 29), tipo="CALL",
       nominal=100000.0):
    return dict(folio=folio, inicio=inicio, venc=venc, tipo=tipo, lado=lado,
                strike=strike, nominal=nominal, contraparte="CP", modalidad="EUROPEA",
                entrega="Compensación", moneda="USD", par="USD/CLP",
                prima_clp=prima_clp, prima_clp_usd=prima_clp_usd,
                prima_pct=prima_pct, prima_usd=prima_usd,
                margen_clp=margen_clp, margen_usd=margen_usd, usdobs=900.0,
                itm_src=None)


# ---------- clasificacion de cartera ----------
def test_clasificar():
    D = date(2026, 5, 29)
    ops = [op("VIG", 900, "Venta", inicio=date(2026, 5, 1), venc=date(2026, 6, 30)),
           op("ALTA", 900, "Venta", inicio=D, venc=date(2026, 6, 30)),
           op("VENCE", 900, "Venta", inicio=date(2026, 4, 1), venc=D),
           op("VIEJA", 900, "Venta", inicio=date(2026, 1, 1), venc=date(2026, 2, 1))]
    vig, ven, altas = dp.clasificar(ops, D)
    assert {o["folio"] for o in vig} == {"VIG", "ALTA"}
    assert [o["folio"] for o in ven] == ["VENCE"]
    assert [o["folio"] for o in altas] == ["ALTA"]


# ---------- deteccion de prima % ----------
def test_es_prima_pct():
    assert dp._es_prima_pct(op("A", 900, "Venta", prima_usd=1500))
    assert not dp._es_prima_pct(op("B", 900, "Venta", prima_clp=1000000))


# ---------- fx por calce ----------
def test_fx_por_calce_1a1():
    ops = [op("V", 920, "Venta", prima_usd=1500, margen_clp=0, margen_usd=0),
           op("C", 920, "Compra", prima_usd=-601.6,
              margen_clp=804993.352, margen_usd=898.4)]
    fx = dp._fx_por_calce(ops)
    k = dp._clave_calce(ops[0])
    assert fx[k] == pytest.approx(896.03, abs=0.01)   # 804993.352/898.4

def test_fx_por_calce_grupos_independientes():
    ops = [op("V1", 920, "Venta", prima_usd=1500, margen_clp=0, margen_usd=0),
           op("C1", 920, "Compra", prima_usd=-601.6, margen_clp=804993.352, margen_usd=898.4),
           op("V2", 930, "Venta", prima_usd=700, margen_clp=0, margen_usd=0),
           op("C2", 930, "Compra", prima_usd=-21.59, margen_clp=607875.7123, margen_usd=678.41)]
    fx = dp._fx_por_calce(ops)
    assert len(fx) == 2                                # un fx por calce
    k1 = dp._clave_calce(ops[0]); k2 = dp._clave_calce(ops[2])
    assert fx[k1] == pytest.approx(804993.352 / 898.4)
    assert fx[k2] == pytest.approx(607875.7123 / 678.41)

def test_fx_por_calce_n_a_m():
    # 2 compras + 1 venta calzadas: agrega margenes y primas del grupo completo
    ops = [op("V", 920, "Venta", prima_usd=2000, margen_clp=0, margen_usd=0),
           op("C1", 920, "Compra", prima_usd=-500, margen_clp=450000, margen_usd=500),
           op("C2", 920, "Compra", prima_usd=-600, margen_clp=360000, margen_usd=400)]
    fx = dp._fx_por_calce(ops)
    k = dp._clave_calce(ops[0])
    assert fx[k] == pytest.approx((450000 + 360000) / (2000 - 500 - 600))


# ---------- P&L segun tipo de prima ----------
def test_pl_prima_clp():
    o = op("N", 900, "Venta", prima_clp=1000000)
    res = [{"mtm": -300000.0, "spot": 900.0}]
    dp._enriquecer_pl([o], res)
    assert res[0]["pl_clp"] == pytest.approx(700000.0)
    assert res[0]["pl_usd"] is None

def test_pl_prima_pct_usa_fx_del_calce():
    ops = [op("V", 920, "Venta", prima_usd=1500, margen_clp=0, margen_usd=0),
           op("C", 920, "Compra", prima_usd=-601.6,
              margen_clp=804993.352, margen_usd=898.4)]
    res = [{"mtm": -90000.0, "spot": 900.0}, {"mtm": 90000.0, "spot": 900.0}]
    dp._enriquecer_pl(ops, res)
    fx = 804993.352 / 898.4
    assert res[0]["pl_usd"] == pytest.approx(1500 - 100)          # V + mtm/spot
    assert res[0]["pl_clp"] == pytest.approx((1500 - 100) * fx)
    assert res[0]["fx_prima"] == pytest.approx(fx)


# ---------- salida de vencimientos: 1 o 2 tablas ----------
def _liq(o):
    return liquidar_vencimiento(dp._op_obj(o), o["usdobs"])

def test_vencimientos_una_tabla(tmp_path):
    ops = [op("N1", 900, "Venta", prima_clp=1000000, prima_clp_usd=10.0)]
    f = tmp_path / "v.xlsx"
    dp.escribir_vencimientos(ops, [_liq(o) for o in ops], str(f), date(2026, 5, 29))
    ws = openpyxl.load_workbook(f).active
    headers = [r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "Folio"]
    assert len(headers) == 1

def test_vencimientos_dos_tablas_y_nota(tmp_path):
    ops = [op("N1", 900, "Venta", prima_clp=1000000, prima_clp_usd=10.0),
           op("P1", 920, "Venta", prima_usd=1500, margen_clp=0, margen_usd=0),
           op("P2", 920, "Compra", prima_usd=-601.6,
              margen_clp=804993.352, margen_usd=898.4)]
    f = tmp_path / "v.xlsx"
    dp.escribir_vencimientos(ops, [_liq(o) for o in ops], str(f), date(2026, 5, 29))
    ws = openpyxl.load_workbook(f).active
    headers = [r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "Folio"]
    assert len(headers) == 2                       # dos tablas
    notas = [ws.cell(r, 19).comment for r in range(1, ws.max_row + 1)
             if ws.cell(r, 19).comment]
    assert len(notas) == 2                         # nota en ambas patas %
    assert "896.03" in notas[0].text
